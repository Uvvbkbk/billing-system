"""
商用对账单生成应用 - Flask 后端 (v3)
支持多客户管理、单位字段、按时间段导出
"""

from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import sqlite3
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import io

app = Flask(__name__)
CORS(app)

# 数据库配置
DATABASE = 'reconciliation.db'

def get_db():
    """获取数据库连接"""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    """初始化数据库"""
    conn = get_db()
    cursor = conn.cursor()
    
    # 创建客户表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # 创建对账单项目表（新结构，包含客户关联和单位）
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS reconciliation_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id INTEGER NOT NULL,
            transaction_date TEXT NOT NULL,
            product_name TEXT NOT NULL,
            quantity REAL NOT NULL,
            unit TEXT,
            unit_price REAL NOT NULL,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (customer_id) REFERENCES customers (id)
        )
    ''')
    
    conn.commit()
    conn.close()

# 初始化数据库
init_db()

@app.route('/')
def index():
    """主页"""
    return render_template('index.html')

# ============ 客户管理 API ============

@app.route('/api/customers', methods=['GET'])
def get_customers():
    """获取所有客户"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM customers ORDER BY name')
    customers = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(customers)

@app.route('/api/customers', methods=['POST'])
def add_customer():
    """添加客户"""
    data = request.json
    customer_name = data.get('name', '').strip()
    
    if not customer_name:
        return jsonify({'error': '客户名称不能为空'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute('INSERT INTO customers (name) VALUES (?)', (customer_name,))
        conn.commit()
        customer_id = cursor.lastrowid
        conn.close()
        return jsonify({'id': customer_id, 'name': customer_name, 'status': 'success'}), 201
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'error': '客户已存在'}), 400

@app.route('/api/customers/<int:customer_id>', methods=['DELETE'])
def delete_customer(customer_id):
    """删除客户及其所有数据"""
    conn = get_db()
    cursor = conn.cursor()
    
    # 删除该客户的所有项目
    cursor.execute('DELETE FROM reconciliation_items WHERE customer_id = ?', (customer_id,))
    # 删除客户
    cursor.execute('DELETE FROM customers WHERE id = ?', (customer_id,))
    
    conn.commit()
    conn.close()
    
    return jsonify({'status': 'success'})

# ============ 对账项目 API ============

@app.route('/api/items', methods=['GET'])
def get_items():
    """获取指定客户的所有对账项目，按日期分组"""
    customer_id = request.args.get('customer_id')
    
    if not customer_id:
        return jsonify({}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT * FROM reconciliation_items 
        WHERE customer_id = ? 
        ORDER BY transaction_date DESC, id DESC
    ''', (customer_id,))
    items = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    # 按日期分组
    grouped = {}
    for item in items:
        date = item['transaction_date']
        if date not in grouped:
            grouped[date] = []
        grouped[date].append(item)
    
    return jsonify(grouped)

@app.route('/api/items', methods=['POST'])
def add_item():
    """添加对账项目"""
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        INSERT INTO reconciliation_items 
        (customer_id, transaction_date, product_name, quantity, unit, unit_price, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (
        int(data.get('customer_id')),
        data.get('transaction_date'),
        data.get('product_name'),
        float(data.get('quantity', 0)),
        data.get('unit', ''),
        float(data.get('unit_price', 0)),
        data.get('notes', '')
    ))
    
    conn.commit()
    item_id = cursor.lastrowid
    conn.close()
    
    return jsonify({'id': item_id, 'status': 'success'}), 201

@app.route('/api/items/<int:item_id>', methods=['PUT'])
def update_item(item_id):
    """更新对账项目"""
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        UPDATE reconciliation_items 
        SET product_name = ?, quantity = ?, unit = ?, unit_price = ?, notes = ?
        WHERE id = ?
    ''', (
        data.get('product_name'),
        float(data.get('quantity', 0)),
        data.get('unit', ''),
        float(data.get('unit_price', 0)),
        data.get('notes', ''),
        item_id
    ))
    
    conn.commit()
    conn.close()
    
    return jsonify({'status': 'success'})

@app.route('/api/items/<int:item_id>', methods=['DELETE'])
def delete_item(item_id):
    """删除对账项目"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM reconciliation_items WHERE id = ?', (item_id,))
    conn.commit()
    conn.close()
    
    return jsonify({'status': 'success'})

@app.route('/api/stats', methods=['GET'])
def get_stats():
    """获取对账统计信息"""
    customer_id = request.args.get('customer_id')
    
    if not customer_id:
        return jsonify({
            'total_items': 0,
            'date_count': 0,
            'total_amount': 0
        })
    
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT COUNT(*) as total FROM reconciliation_items WHERE customer_id = ?', (customer_id,))
    total = cursor.fetchone()['total']
    
    cursor.execute('SELECT COUNT(DISTINCT transaction_date) as date_count FROM reconciliation_items WHERE customer_id = ?', (customer_id,))
    date_count = cursor.fetchone()['date_count']
    
    cursor.execute('SELECT SUM(quantity * unit_price) as total_amount FROM reconciliation_items WHERE customer_id = ?', (customer_id,))
    total_amount = cursor.fetchone()['total_amount'] or 0
    
    conn.close()
    
    return jsonify({
        'total_items': total,
        'date_count': date_count,
        'total_amount': round(total_amount, 2)
    })

@app.route('/api/export', methods=['POST'])
def export_reconciliation():
    """导出对账单为 Excel（支持时间段筛选）"""
    data = request.json
    customer_id = data.get('customer_id')
    start_date = data.get('start_date')
    end_date = data.get('end_date')
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 获取客户名称
    cursor.execute('SELECT name FROM customers WHERE id = ?', (customer_id,))
    customer_row = cursor.fetchone()
    customer_name = customer_row['name'] if customer_row else '未知客户'
    
    # 查询指定时间段内的数据
    if start_date and end_date:
        cursor.execute('''
            SELECT * FROM reconciliation_items 
            WHERE customer_id = ? AND transaction_date >= ? AND transaction_date <= ?
            ORDER BY transaction_date, id
        ''', (customer_id, start_date, end_date))
    else:
        cursor.execute('''
            SELECT * FROM reconciliation_items 
            WHERE customer_id = ?
            ORDER BY transaction_date, id
        ''', (customer_id,))
    
    items = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    # 按日期分组
    grouped_by_date = {}
    for item in items:
        date = item['transaction_date']
        if date not in grouped_by_date:
            grouped_by_date[date] = []
        grouped_by_date[date].append(item)
    
    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = '对账明细'
    
    # 设置列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    
    # 定义样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    date_header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    date_header_font = Font(bold=True, color='FFFFFF', size=10)
    subtotal_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    subtotal_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')
    
    # 添加标题
    title = f"{customer_name}&泽宇辅料销售对账单"
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = center_alignment
    
    # 添加电话信息（居中）
    ws['A2'] = "TEL：14705991833"
    ws.merge_cells('A2:G2')
    ws['A2'].alignment = center_alignment
    
    current_row = 4
    
    # 按日期遍历
    for date in sorted(grouped_by_date.keys()):
        items_for_date = grouped_by_date[date]
        
        # 添加日期行
        ws.cell(row=current_row, column=1).value = f"日期: {date}"
        ws.merge_cells(f'A{current_row}:G{current_row}')
        ws.cell(row=current_row, column=1).fill = date_header_fill
        ws.cell(row=current_row, column=1).font = date_header_font
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1
        
        # 添加表头
        headers = ['序号', '品名', '数量', '单位', '单价', '金额', '备注']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        current_row += 1
        
        # 添加该日期的数据行
        date_total_amount = 0
        for idx, item in enumerate(items_for_date, 1):
            amount = item['quantity'] * item['unit_price']
            date_total_amount += amount
            
            ws.cell(row=current_row, column=1).value = idx
            ws.cell(row=current_row, column=2).value = item['product_name']
            ws.cell(row=current_row, column=3).value = int(item['quantity'])
            ws.cell(row=current_row, column=4).value = item['unit']
            ws.cell(row=current_row, column=5).value = item['unit_price']
            ws.cell(row=current_row, column=6).value = amount
            ws.cell(row=current_row, column=7).value = item['notes']
            
            # 应用边框和对齐
            for col in range(1, 8):
                cell = ws.cell(row=current_row, column=col)
                cell.border = border
                if col == 1:
                    cell.alignment = Alignment(horizontal='left')
                elif col == 3:
                    cell.number_format = '0'
                    cell.alignment = right_alignment
                elif col in [5, 6]:
                    cell.number_format = '0.00'
                    cell.alignment = right_alignment
                else:
                    cell.alignment = Alignment(horizontal='left')
            
            current_row += 1
        
        # 添加小计行
        ws.cell(row=current_row, column=1).value = '小计'
        ws.cell(row=current_row, column=1).fill = subtotal_fill
        ws.cell(row=current_row, column=1).font = subtotal_font
        ws.cell(row=current_row, column=1).border = border
        
        ws.cell(row=current_row, column=6).value = date_total_amount
        ws.cell(row=current_row, column=6).fill = subtotal_fill
        ws.cell(row=current_row, column=6).font = subtotal_font
        ws.cell(row=current_row, column=6).number_format = '0.00'
        ws.cell(row=current_row, column=6).alignment = right_alignment
        ws.cell(row=current_row, column=6).border = border
        
        for col in [2, 3, 4, 5, 7]:
            ws.cell(row=current_row, column=col).fill = subtotal_fill
            ws.cell(row=current_row, column=col).border = border
        
        current_row += 2  # 空一行
    
    # 添加总计行
    ws.cell(row=current_row, column=1).value = '总计'
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=current_row, column=1).border = border
    
    total_amount = sum(item['quantity'] * item['unit_price'] for item in items)
    ws.cell(row=current_row, column=6).value = total_amount
    ws.cell(row=current_row, column=6).font = Font(bold=True, size=11)
    ws.cell(row=current_row, column=6).number_format = '0.00'
    ws.cell(row=current_row, column=6).alignment = right_alignment
    ws.cell(row=current_row, column=6).border = border
    
    for col in [2, 3, 4, 5, 7]:
        ws.cell(row=current_row, column=col).border = border
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 生成文件名：客户名+对账单+结束日期
    end_date_str = end_date if end_date else datetime.now().strftime('%Y-%m-%d')
    end_date_formatted = end_date_str.replace('-', '')
    filename = f"{customer_name}对账单{end_date_formatted}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
